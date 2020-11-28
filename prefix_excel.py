"""人口動態分析のために変則的なエクセルファイルを前処理するスクリプト。
ひとつのブックに数か月分（ふつうは12ヵ月分）のシートが含まれている場合に使用する。
ブックに含まれる複数のシートを個別のブックに分解して保存する。
ブックのタイプの自動判定は行えない"""

import tkinter
from tkinter import filedialog
import pandas
import openpyxl
import os
import re
import datetime
from dateutil.relativedelta import relativedelta
from set_pref import *
import globalize
import prefix_two_columns
import moji_code


# 作成したファイルを保存するパスを決めるクラス
class ExcellPath:

    def write_excell(self, prefa, sheetdata):
        for data in sheetdata:
            df = data[0]
            sheetname = data[1]
            sheetname = sheetname.replace(".", "_")
            path = "C:/Users/lemat/lempy/poll/data" + "/" + prefa + "/" + sheetname + ".xlsx"
            dirpath = "C:/Users/lemat/lempy/poll/data" + "/" + prefa
            os.makedirs(dirpath, exist_ok=True)
            df.to_excel(path)


# シートを別々のブックに分解するクラス
class SeparateSheets:
    sheetdata = []

    def read_excel(self):
        typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        for p in path:
            book = pandas.ExcelFile(p)
            list_of_sheetnames = book.sheet_names
            for sheet in list_of_sheetnames:
                df = pandas.read_excel(p, sheet_name=sheet, dtype="object")
                self.sheetdata.append([df, sheet])


    def read_csv(self):
        typ = [("CSV", "csv")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        for p in path:
            c_code = moji_code.check_encoding(p)[2]
            df = pandas.read_csv(p, dtype="object", encoding=c_code)
            self.sheetdata.append([df]) #ｃｓｖの場合、リストにシート名を入れないので注意


# エクセルファイルを修正する
class ModifyExel:
    sheetdata = []

    def osaka_excel2(self): # osaka_excel1はないので注意のこと
        print("このデータは分析用に整形して、直接/data/osakaに出力します。")
        typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        list_of_dfs = []
        for p in path:
            book = pandas.ExcelFile(p)
            list_of_sheetnames = book.sheet_names
            year = re.search(r'...(?=c)', p).group().zfill(3)  # 和暦から年データをつくる
            if year[0] == "h":
                year = int(year[1: 3]) + 1989
            elif year[1] == "r":
                year = int(year[1: 3]) + 2018
            for sheet in list_of_sheetnames:
                df = pandas.read_excel(p, sheet_name=sheet, dtype="object")
                # 切り出し位置を決める
                df2 = df.rename(columns={"Unnamed: 68": "new0"})
                # 上半分の位置
                dfc = df2[df2["new0"].str.contains("PAGE －   1", na=False)]
                start_nc = dfc.index.values[0]
                dfc2 = df2[df2["new0"].str.contains("PAGE －   2", na=False)]
                end_nc = dfc2.index.values[0] - 1
                # 下半分の位置
                start_nd = dfc2.index.values[0]
                end_nd = start_nd + (end_nc - start_nc)
                # インデックスをリセット。空白列をカット。
                dfc3 = df2.iloc[:0]
                pc = dfc3.columns.values[2]
                df = df.rename(columns={pc: "place"})
                df = df.set_index("place")
                df.dropna(how="all", axis=1, inplace=True)
                # 2段になっているので切り出す
                dfa = df[start_nc: end_nc].copy()
                dfb = df[start_nd: end_nd].copy()
                temp_l =[dfa, dfb]
                for dfx in temp_l:
                    dfx = dfx.drop(dfx.index[[0]])
                    dfx = dfx.dropna(how="all", axis=0)
                    dfx = dfx.dropna(how="all", axis=1)
                    #月名のある行を取り出し、月リストを作り、必要な列の番号を取得する
                    d_d = dfx.iloc[:1]
                    temp_ml = d_d.values.tolist() #このリストは二次元なので注意すること！
                    for i, element in enumerate(temp_ml[0]):
                        if "月" in str(element):
                            # 月リストをつくる
                            element = str(element).replace("\u3000", "")
                            # 数字を半角に直す
                            zen = "０１２３４５６７８９"
                            han = "0123456789"
                            trans_table = str.maketrans(zen, han)
                            element = element.translate(trans_table)
                            # 次ごとのデータフレームを切り出す
                            dfm = dfx.iloc[:, [i]].copy()
                            # date欄を作る
                            month = re.search(r'\d{1,2}(?=月)', element).group().zfill(2)  # 数字だけ取り出して二桁にする
                            date = "{}-{}-28".format(year, month)
                            dfm["date"] = date
                            # 日付を月初にする
                            dfm["date"] = pandas.to_datetime(dfm["date"])  # date列を日付データとして使う
                            f_nextmonth = lambda x: x + relativedelta(months=+1)  # 月を一月後にする（同月のデータのため）
                            dfm["date"] = dfm["date"].apply(f_nextmonth)
                            f_firstday = lambda x: x + relativedelta(day=1)  # 日付を１日にする（はっきりさせるため）
                            dfm["date"] = dfm["date"].apply(f_firstday)
                            # todo place列の総数を大阪府に
                            dfm = dfm.rename(index={"総数": "大阪府"})
                            dfm.columns = ["death", "date"]
                            dfm.insert(0, "total", 0)
                            dfm = dfm.reset_index()
                            list_of_dfs.append(dfm)
        return(list_of_dfs)

    def kyoto_excel(self):
        print("2015年10月の.csvファイルのdammy列5行目に「総数」と手動で書き込む必要があります。")
        typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        for p in path:
            book = pandas.ExcelFile(p)
            list_of_sheetnames = book.sheet_names
            for sheet in list_of_sheetnames:
                df = pandas.read_excel(p, sheet_name=sheet, dtype="object")
                # 年月データをまとめる
                #print(df.iat[1, 0], df.iat[1, 1])
                df.iat[0, 0] = df.iat[1, 0] + df.iat[1, 1]
                df = df.drop(1)
                # 年月日欄を明示
                df.iat[1,0] = "市町村"
                # 0列名に市町村をそろえる
                dfc = df.iloc[37:48, :].copy()
                dfc.iloc[:, 0] = dfc.iloc[:, 1]
                df = pandas.concat([df, dfc])
                if sheet == "1510（参考）":
                    df = df.assign(x=0, y=0)
                    df = df.rename(columns= {"x": "Unnamed: 20", "y": "Unnamed: 21"})
                    df.insert(2, "dammy", 0)
                    df.dammy = df.dammy.replace(5, "総数") #ファイルに反映されない
                tempdata = []
                tempdata.append(df)
                tempdata.append(sheet)
                self.sheetdata.append(tempdata)


    def tokyo_excel(self):
        typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        # 開いたブックを処理
        for p in path:
            book = pandas.ExcelFile(p)
            list_of_sheetnames = book.sheet_names
            # 開いたシートを処理
            for sheet in list_of_sheetnames:
                df = pandas.read_excel(p, sheet_name=sheet, dtype="object")
                print(p)
                #切り出す位置を決める
                df2 = df.rename(columns={"Unnamed: 0": "new0"})
                dfc = df2[df2["new0"].str.contains("人口の動き", na=False)]
                print(dfc.iloc[0, 0])
                start_n = dfc.index.values[0]
                dfd = df2[df2["new0"].str.contains("及び外国人の合計である。", na=False)]
                end_n = dfd.index.values[0]
                #必要な表を切り出す
                dfs = df[start_n: end_n].copy()
                # 日付を取得して書き加える
                # dfの10行目までをリストにする
                search_date = df[0:10].values.tolist()
                #日付を取得して、dfsに書き加える
                for get_date in search_date:
                    for s in get_date:
                        if "現在" in str(s):
                            nen_getsu = s
                            break
                nen_getsu = nen_getsu.strip()
                dfs.iloc[0, 0] = nen_getsu
                # キーワードを書き込む
                dfs.iloc[2, 1] = "市町村"
                dfs.iloc[2, 2] = "総数"
                dfs.iloc[4, 0] = "削除"
                # データフレームをリストに入れる
                tempdata = []
                tempdata.append(dfs)
                tempdata.append(nen_getsu)
                self.sheetdata.append(tempdata)


    def tokyo_excel2(self):
        print("このデータは分析用に整形して、直接/data/tokyoに出力します。")
        typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        list_of_dfs = []
        for p in path:
            book = pandas.ExcelFile(p)
            list_of_sheetnames = book.sheet_names
            for sheet in list_of_sheetnames:
                df = pandas.read_excel(p, sheet_name=sheet, dtype="object")
                df.rename(index={0: "n1", 1: "n2"}, inplace=True)
                year = re.search(r'(?<=ju)\d{2}', p).group()
                year = "20" + year
                list_of_contents = df.values.tolist() #1行目と2行目を取り出してスペースを消す
                l_c0 = []
                l_c1 = []
                for element in list_of_contents[0]:
                    element = str(element).replace("\u3000", "")
                    l_c0.append(element)
                for element in list_of_contents[1]:
                    element = str(element).replace("\u3000", "")
                    l_c1.append(element)
                l_c = [l_c0, l_c1] #元のデータフレームに１行目と２行目を戻す
                dfi = pandas.DataFrame(l_c)
                l_i= dfi.columns.values
                df.columns = l_i
                dfn = pandas.concat([dfi, df], axis=0)
                dfn.drop(["n1", "n2"], inplace=True)
                for i in range(2, 14): #月ごとの表を作って日付列を追加する
                    dfa = dfn.iloc[:, [1, i]].copy()
                    month = dfa.iat[0, 1]
                    # 数字を半角に直す
                    zen = "０１２３４５６７８９"
                    han = "0123456789"
                    trans_table = str.maketrans(zen, han)
                    month = month.translate(trans_table)
                    month = re.search(r'\d{1,2}(?=月)', month).group().zfill(2) #数字だけ取り出して二桁にする
                    # データフレームに日付を加える
                    date = "{}-{}-28".format(str(year), month)
                    dfa["date"] = date
                    # 日付を月初にする
                    dfa["date"] = pandas.to_datetime(dfa["date"])  # date列を日付データとして使う
                    f_nextmonth = lambda x: x + relativedelta(months=+1)  # 月を一月後にする（同月のデータのため）
                    dfa["date"] = dfa["date"].apply(f_nextmonth)
                    f_firstday = lambda x: x + relativedelta(day=1)  # 日付を１日にする（はっきりさせるため）
                    dfa["date"] = dfa["date"].apply(f_firstday)
                    dfa.drop([0, 1, 2], inplace=True)
                    dfa.at[3, 1] = "東京都"
                    dfa.at[5, 1] = "東京都区部"
                    dfa.columns = ["place", "death", "date"]
                    dfa.insert(1, "total", 0)
                    list_of_dfs.append(dfa)
        return(list_of_dfs)



    def chiba_excel(self, sheetdata):
        df = sheetdata[0]
        if "日本人" in sheetdata[1] or "外国人" in sheetdata[1]:
            return
        # 年月データを含む要素の座標を抽出する
        date_loc = [] # 年月データの座標
        c = len(df.columns)
        for i in range(c):
            df_check = df.iloc[:, i]
            d = len(df_check)
            for j in range(d):
                if "月中" in str(df_check[j]):
                    date_loc =[j, i]
        rn = date_loc[0]
        cn = date_loc[1]
        df.iloc[rn, cn] = df.iloc[rn, cn].replace("月中", "月28日")
        df.iloc[rn, cn] = df.iloc[rn, cn].strip()
        df.iloc[1, 1] = "市町村"
        df.iloc[1, 2] = "総数"
        sheetname = df.iloc[rn, cn]
        print(sheetname + "をファイル名にします")
        tempdata = []
        tempdata.append(df)
        tempdata.append(sheetname)
        self.sheetdata.append(tempdata)

    def kanagawa_excel(self, sheetdata):
        df = sheetdata[0]
        date_array = df.columns.values
        #和暦を整形する
        for element in date_array:
            if "月  中" in element or "月中" in element:
                date_index = element
                date = date_index.replace(" ", "")
                date = date.replace("　", "")
                # 全角数字を半角数字にする
                zen = "０１２３４５６７８９"
                han = "0123456789"
                trans_table = str.maketrans(zen, han)
                date = date.translate(trans_table)
                # 直後に「年」がある「元」を１に置換する
                if "元" in date:
                    date = re.sub(r'元(?=年)', "1", date)
                # 日付を28日にする
                date = date.replace("月中", "月28日")
                # 日付表示を書き換える
                df = df.rename(columns={date_index: date})
        # 適切な列に「総数」という文字を入れる
        df.iloc[1, 4] = "総数"
        sheetname = date
        print(sheetname + "をファイル名にします")
        tempdata = []
        tempdata.append(df)
        tempdata.append(sheetname)
        self.sheetdata.append(tempdata)

    def saitama_excel(self, sheetdata):
        #シートを判別する
        df = sheetdata[0]
        date_array = df.columns.values
        f = 0
        for title in date_array:
            if "日現在" in title:
                f = 1
        if f == 1:
            # 和暦を整形する
            for element in date_array:
                if "日現在" in element:
                    date_index = element
                    date = date_index.replace("　", "")
                    date = date.replace("  ", "")
                    df = df.rename(columns={date_index: date})
            if df.iat[1, 10] == "死":
                df.iloc[1, 10] = "死亡"
            elif df.iat[1, 11] == "死":
                df.iloc[1, 11] = "死亡"
            else:
                print("死亡欄の修正はおこなっていません。")
            sheetname = date
            print(sheetname + "をファイル名にします")
            tempdata = []
            tempdata.append(df)
            tempdata.append(sheetname)
            self.sheetdata.append(tempdata)
        else:
            return


    def aichi_file(self, sheetdata):
        df = sheetdata[0]
        # 日付を整形
        date_d = df.loc[0, "期間"]
        year = int(date_d[0:4])
        month = date_d[4:6]
        day = 28
        # day = date_d[9:11]
        #read_excelで使うために和暦を挿入しておく
        if year < 2019:
            nen = year - 1988
            nen = "平成{}".format(nen)
        elif year > 2018:
            nen = year - 2018
            nen = "令和{}".format(nen)
        else:
            nen = year - 1925
            nen = "昭和{}".format(nen)
        nengappi = "{}年{}月{}日現在".format(nen, month, day)
        df = df.rename(columns={"県市区町村": "自治体", "増減数": "総数", "期間": nengappi})
        # 「日外他」「男女」「総数」を切り出す
        df_all = df.query("国籍区分 == '日外他' and 性別 == '男女' and 年齢区分 == '総数'")
        #　年齢区分列を削除
        df_all.drop(columns="年齢区分", inplace=True)
        # 死亡のマイナスを外す
        df_all["死亡"] = df_all["死亡"].astype(int).copy()
        f_reverse = lambda x: -x
        df_all["死亡"] = df_all["死亡"].map(f_reverse)  # 死亡列をseriesとして取り出してmapでラムダ関数を適用
        sheetname = "{}{}".format(year, month)  # ファイル名をつける
        # 1行目に列名を入れる
        c_index = df_all.columns.values
        df_c = pandas.DataFrame(c_index)
        df_c = df_c.transpose()
        df_c.columns = c_index
        df_a = pandas.concat([df_c, df_all])
        sheetname = "{}{}".format(year, month) #ファイル名をつける
        print(sheetname + "をファイル名にします")
        tempdata = []
        tempdata.append(df_a)
        tempdata.append(sheetname)
        self.sheetdata.append(tempdata)


    def aichi_excel2(self):
        print("このデータは分析用に整形して、直接/data/aichiに出力します。")
        typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
        path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
        list_of_dfs = []
        for p in path:
            book = pandas.ExcelFile(p)
            list_of_sheetnames = book.sheet_names
            df = pandas.read_excel(p, sheet_name=list_of_sheetnames[0], dtype="object")
            # 和暦から年データを取り出す
            d_y = df.iloc[0:4]
            l_y = d_y.values.tolist()
            for array in l_y:
                for element in array:
                   if "年" in str(element):
                        year = element
                        break
            #各シートから月別のデータフレームを切り出す
            for sheet in list_of_sheetnames:
                #保健所列を探す
                df = pandas.read_excel(p, sheet_name=sheet, dtype="object")
                d_i = df.iloc[0:4]
                l_i = d_i.values.tolist()
                loc_of_place = []
                for array in l_i:
                    for i, element in enumerate(array):
                        if "保健所" in str(element):
                            loc_of_place.append(i)  # 存在しない行のデータは不要なのでこれで
                #政令指定都市、中核市の都市名をひとつ右のセルに移す
                d_bc = df.iloc[:, loc_of_place[0]+1]
                bc = d_bc.values.tolist()
                for j, ci in enumerate(bc):
                    ci = str(ci)
                    if "愛知県" in ci or "名古屋市" in ci or "豊橋市" in ci or "豊田市" in ci:
                        df.iloc[j, loc_of_place[0]+2] = df.iloc[j, loc_of_place[0]+1]
                # 二つめ以降の保健所列、市町村列を消す
                for pc2 in loc_of_place[1:]:
                    df.drop(df.columns[[pc2, pc2+1, pc2+2]], axis=1)
                # 一つめの保健所列と周辺の空白列を消す。市町村列をインデックスにする
                pc = df.columns.values[loc_of_place[0] + 2]
                df = df.rename(columns={pc: "place"})
                df = df.drop(df.columns[[loc_of_place[0], loc_of_place[0]+1, loc_of_place[0]+3]], axis=1)
                df = df.set_index("place")
                # 月ごとのデータフレームに分解し、date, total（ダミー）を加える
                d_d = df.iloc[0:4]
                l_d = d_d.values.tolist()
                for array in l_d:
                    for i, element in enumerate(array):
                        if "月" in str(element):
                            month = element
                            month = month.replace("　", "")
                            date = "{}{}28日".format(year,month)
                            date = globalize.wareki(date)
                            dfm = df.iloc[:, [i]]
                            dfm.insert(0, "total", 0)
                            dfm.insert(2, "date", date)
                            dc = dfm.columns.values[1]
                            dfm = dfm.rename(columns={dc: "death"})
                            # placeを列に戻す
                            dfm = dfm.reset_index()
                            pandas.set_option("display.max_rows", None)
                            list_of_dfs.append(dfm)
        return(list_of_dfs)


# エクセルファイルを開く
def open_excel():
    typ = [("エクセル", "xlsx"), ("エクセル", "xls")]
    path = tkinter.filedialog.askopenfilename(initialdir="./", filetypes=typ, multiple=True)
    return(path)

def hyogo():
    pref = Hyogo()
    print(pref.pref)
    open = SeparateSheets()
    print("ファイルを処理します。")
    open.read_excel()
    write = ExcellPath()
    print("ファイルを書き込んでいます。")
    write.write_excell(pref.prefa, open.sheetdata)

def osaka():
    pref = Osaka()
    print(pref.pref)
    # path = open_excel()
    open = SeparateSheets()
    print("ファイルを処理します。")
    open.read_excel()
    write = ExcellPath()
    print("ファイルを書き込んでいます。")
    write.write_excell(pref.prefa, open.sheetdata)

def osaka2():
    pref = Osaka()
    print(pref.pref, "タイプ２")
    open = ModifyExel()
    list_of_dfs = open.osaka_excel2()
    prefix_two_columns.main(list_of_dfs, pref.pref, pref.prefa)


def kyoto():
    pref = Kyoto()
    print(pref.pref)
    open = ModifyExel()
    open.kyoto_excel()
    write = ExcellPath()
    print("ファイルを書き込んでいます。")
    write.write_excell(pref.prefa, open.sheetdata)

def tokyo():
    pref = Tokyo()
    print(pref.pref)
    open = ModifyExel()
    open.tokyo_excel()
    write = ExcellPath()
    print("ファイルを書き込んでいます。")
    write.write_excell(pref.prefa, open.sheetdata)

def tokyo2(): #年単位でまとまっている月別データを処理する
    pref = Tokyo()
    print(pref.pref, "タイプ２")
    open = ModifyExel()
    list_of_dfs = []
    list_of_dfs = open.tokyo_excel2()
    prefix_two_columns.main(list_of_dfs, pref.pref, pref.prefa)

def chiba():
    pref = Chiba()
    print(pref.pref) # 注意喚起用
    open = SeparateSheets()
    print("ファイルを処理します。")
    open.read_excel()
    for sheet in open.sheetdata:
        o2 = ModifyExel()
        o2.chiba_excel(sheet)
    print("ファイルを書き込んでいます。")
    write = ExcellPath()
    write.write_excell(pref.prefa, o2.sheetdata)


def kanagawa():
    pref = Kanagawa()
    print(pref.pref)
    open = SeparateSheets()
    print("ファイルを処理します。")
    open.read_excel()
    for sheet in open.sheetdata:
        o2 = ModifyExel()
        o2.kanagawa_excel(sheet)
    print("ファイルを書き込んでいます。")
    write = ExcellPath()
    write.write_excell(pref.prefa, o2.sheetdata)


def saitama():
    pref = Saitama()
    print(pref.pref)
    print("2007(平成17）年2月のファイルに原因不明のエラーがあります。手動で修正してください。")
    print("2007（平成19）年9月分以降分には、このスクリプトによる処理は必要ありません。")
    open = SeparateSheets()
    print("ファイルを処理します。")
    open.read_excel()
    for sheet in open.sheetdata:
        o2 = ModifyExel()
        o2.saitama_excel(sheet)
    print("ファイルを書き込んでいます。")
    write = ExcellPath()
    write.write_excell(pref.prefa, o2.sheetdata)


def aichi():
    pref = Aichi()
    print(pref.pref)
    open = SeparateSheets()
    print("ファイルを処理します。")
    open.read_csv()
    for sheet in open.sheetdata:
        o2 = ModifyExel()
        o2.aichi_file(sheet) #この処理ラインではシート名は入れていない。関数側でファイル名生成必要
    print("ファイルを書き込んでいます。")
    write = ExcellPath()
    write.write_excell(pref.prefa, o2.sheetdata)


def aichi2():
    pref = Aichi()
    print(pref.pref, "タイプ２")
    open = ModifyExel()
    list_of_dfs = open.aichi_excel2()
    prefix_two_columns.main(list_of_dfs, pref.pref, pref.prefa)

if __name__ == "__main__":
    aichi2()




